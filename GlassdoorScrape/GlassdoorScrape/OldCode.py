#Start custom Python coding
#script to convert GlassDoor Company To IDs. in Python 3.6.4, then harvest each company onto a separate csv list
#It also saves the current page on the Excel Worksheet
#Requirements: Need openpyxl Excel package installed and lxml

#import required by custom script
import datetime
#Alteryx imports
#import AlteryxPythonSDK as Sdk
#import xml.etree.ElementTree as Et

#variables required by script
filenameExcel = "C:\\Users\\nick.hayden\\AppData\\Roaming\\Alteryx\\Tools\\GlassdoorScrapeALPHA\\data\\CompanyIDs.xlsx" #which companies?
ColumnCompanyName = "C"
ColumnGlassDoorID = "I"
ColumnGlassDoorCompanyName = "J"
ColumnGlassDoorFullLink = "K"
ColumnPageNo = "L"
RowStartData = 2
fileobject=None
#endregion variables required by script


#region userdefined functions, utilities
def DownloadString(URL):

    try:
        from urllib.request import Request, urlopen  # Python 3
    except ImportError:
        from urllib2 import Request, urlopen  # Python 2
    
    html = ""
    while html == "":
            try:
                print("Download string from "+URL)
                q = Request(URL)
                q.add_header('User-Agent', 'Mozilla/5.0')
                charset='utf-8'
                resource=urlopen(q)
                try:
                    html = resource.read().decode(charset)
                except:
                    charset='unicode'
                    html = resource.read().decode(charset)
            except urllib2.HTTPError as err:
                print("Error fetching URL, To retry after 15 seconds. ERROR CODE: "+err.code)
                if(err.code == 404):
                    return None
                PauseScript(15)
    return html


#Pauses script execution
def PauseScript(SecondsPause):
    import time
    print("Sleeping for "+str(SecondsPause)+" seconds")
    time.sleep(SecondsPause)
    

#Writes to a file and returns a fileobject if it was not defined
#If its already defined the new line it appends a new line before writing the new string
def WriteToFile(stringToWrite, fileobject, filepath, NewLine="\r\n"):
    import codecs
    if fileobject is None: #Check if fileobject variable is None?
        fileobject = codecs.open(filepath, 'w', "utf-8")
        fileobject.write(stringToWrite)
    else:
        fileobject.write(NewLine)
        fileobject.write(stringToWrite)
    fileobject.flush()
    return fileobject



#Returns a list of substrings found in between strings
def GetListOfSubstrings(stringSubject,string1,string2,Limit=-1):
    MyList = []
    intstart=0
    strlength=len(stringSubject)
    continueloop = 1
    
    while(intstart < strlength and continueloop == 1 and len(MyList) != Limit):
        intindex1=stringSubject.find(string1,intstart)
        if(intindex1 != -1): #The substring was found, lets proceed
            intindex1 = intindex1+len(string1)
            intindex2 = stringSubject.find(string2,intindex1)
            if(intindex2 != -1):
                subsequence=stringSubject[intindex1:intindex2]
                subsequence = subsequence.strip()
                MyList.append(subsequence)
                intstart=intindex2+len(string2)
            else:
                continueloop=0
        else:
            continueloop=0
    return MyList    
    

#Returns a directory path name
def GetDirectoryPath(fileNamePath):
    import os.path
    return os.path.dirname(fileNamePath)
    

#Returns a string between two tags
def GetStringBetween(stringSubject,string1,string2,ReturnStringIfNotFound="",FormatAsHtml=True):
    MyList=GetListOfSubstrings(stringSubject,string1,string2,1)
    if(len(MyList) == 0):
        return ReturnStringIfNotFound
    TheString=MyList[0]
    if(FormatAsHtml == True):
        return StripHtml(TheString)
    else:
        return TheString


#Strips html from a string, you need to make sure lxml is installed first
def StripHtml(stringSubject):
    from lxml import html
    from lxml.html.clean import clean_html    
    tree = html.fromstring(stringSubject)
    clean_tree = clean_html(tree)
    return clean_tree.text_content().strip()


#converts a list to a csv string
def ConvertToCSVString(List):
    CsvString=""
    for x in range(0, len(List)):
        str = List[x]
        str = str.replace("\r", " ")
        str = str.replace("\n", " ")
        str = str.replace("  ", " ")
        if(str.find("\"", 0) != -1):
            str = str.replace("\"", "\"\"")
        if(str.find(",", 0) != -1):
            str = "\"" + str + "\"" 
        CsvString = CsvString + str
        if(x + 1 < len(List)):
            CsvString = CsvString + ","
    return CsvString

#Loads a workbook, returns a workbook object
def LoadWorkBook(InputFileName):
    # Import `load_workbook` module from `openpyxl`
    from openpyxl import load_workbook

    #Load the Excel file
    wb = load_workbook(InputFileName)
    return wb

#Encodes a string to be URL safe
def EncodeURL(stringToEncode):
    from urllib.parse import quote
    return quote(stringToEncode)

#endregion



#region ScrapeSpecificFunctions
def DecodePageNo(PageNoString):
    try:
        number = int(PageNoString.strip())
        return number
    except:
        return 1


def GetCSVHeader():
    List = []
    List.append("Company Name")
    List.append("Review Date")
    List.append("Helpful (count)")
    List.append("Title (of the review)")
    List.append("Rating (out of 5)")
    List.append("Current/ Past Employee")
    List.append("Employee Title")
    List.append("Employment Type")
    List.append("Location")
    List.append("Recommends")
    List.append("Positive Outlook")
    List.append("Approves of CEO")
    List.append("Time Employed")
    List.append("Pros")
    List.append("Cons")
    List.append("Advice to Management")
    return ConvertToCSVString(List)

#BoxColor
def DecodeBoxColor(HTMLString):
    if(HTMLString.find("yellow", 0) != -1):
        return "Yellow"
    elif(HTMLString.find("green", 0) != -1):
         return "Green"
    elif(HTMLString.find("red", 0) != -1):
         return "Red"
    return ""


#Parses the HTML
def ParseHTML(HTMLString):
    List = GetListOfSubstrings(HTMLString,"<div class='hreview'>","<li class=' empReview")
    print("Reviews to parse: "+str(len(List)))
    #Each will have a list appended
    OutputListing = []
    for mainlist in range(0, len(List)):
        row=List[mainlist]
        #Parse the current listing
        RowList=[]
        #Company Name
        _str = GetStringBetween(HTMLString, "</script><title>", "Reviews |","")
        RowList.append(_str)
        
        #ReviewDate
        _str = GetStringBetween(row, "datetime=\"","\">","")
        RowList.append(_str)
        #Helpful count
        _str = GetStringBetween(row, "<span class=\"helpfulCount subtle\">","</span>","")
        _str = _str.replace("Helpful","")
        _str = _str.replace("(","")
        _str = _str.replace(")","")
        _str = _str.strip()
        RowList.append(_str)
        #Title (of the review)
        _str = GetStringBetween(row, "<span class=\"summary \">","</span>","")
        _str = _str.strip("\"")
        RowList.append(_str)
        #Rating of Review
        _str = GetStringBetween(row, "<span class=\"rating\"><span class=\"value-title\" title=\"","></span>","")
        _str = _str.strip("\"")
        RowList.append(_str)
        #Current Employee / Past
        _str = GetStringBetween(row,"<span class='authorJobTitle middle reviewer'>", "</span>" , "")
        _str = _str.strip()
        _current = _str
        _current = GetStringBetween( (">"+_current), ">", "Employee" , "")
        RowList.append(_current)
        
        #Employee title
        _current = _str
        _current = GetStringBetween(_current+"</", "-", "</" , "")
        RowList.append(_current)
        
        #Employee type
        _str = GetStringBetween(row, "<p class=' tightBot mainText'>","</p>","")
        if(_str.find("full", 0) != -1):
            RowList.append("Full time")
        elif(_str.find("half", 0) != -1):
            RowList.append("Half time")
        else:
            RowList.append(_str)
        

        #Location
        _str = GetStringBetween(row, "<span class='authorLocation middle'>", "></span>" ,"")
        RowList.append(_str)

        #Recommends
        _str = GetStringBetween(row, "<i class" , "Recommends</span>" , "", False)
        _str = DecodeBoxColor(_str)
        RowList.append(_str)

        #Outlook
        _str = GetStringBetween(row, "</i></div><div class=\"cell\"><span class='middle'>" , "Outlook</span>" , "")
        if(_str.find("Positive", 0) != -1):
            RowList.append("Positive")
        elif(_str.find("Negative", 0) != -1):
            RowList.append("Negative")
        elif(_str.find("Neutral", 0) != -1):
            RowList.append("Neutral")
        else:
            RowList.append(_str)

        #CEO
        _str = GetStringBetween(row, "<i class" , "CEO</span>" , "", False)
        _str = DecodeBoxColor(_str)
        RowList.append(_str)

        #Time Employed:
        _str = GetStringBetween(row, "<p class=' tightBot mainText'>","</p>","")
        if(_str.find("More than a year", 0) != -1):
            RowList.append("More than a year")
        elif(_str.find("Less than a year", 0) != -1):
            RowList.append("Less than a year")
        else:
            RowList.append(_str)
        
        #Pros
        _str = GetStringBetween(row, "<p class=' pros mainText truncateThis wrapToggleStr'>", "</p>" ,"")
        RowList.append(_str)

        #Cons
        _str = GetStringBetween(row, "<p class=' cons mainText truncateThis wrapToggleStr'>", "</p>" ,"")
        RowList.append(_str)

        #Advice to management
        _str = GetStringBetween(row, "<p class=' adviceMgmt mainText truncateThis wrapToggleStr'>", "</p>" ,"")
        RowList.append(_str)
            

        #append the list
        OutputListing.append(RowList)
        
    return OutputListing


    
#endregion




#load the filename from panda
print("Loading the Excel file...")
wb = LoadWorkBook(filenameExcel)
ws = wb.active
print("Harvesting file to put into :"+filenameExcel)
colrange = ws[ColumnCompanyName]

EndRange = RowStartData + len(colrange)

for rows in range(RowStartData, EndRange):
    #Retrieve the cell values
    stringCompanyName = ws[ColumnCompanyName + str(rows)].value
    companyNameFromSite = ws[ColumnGlassDoorCompanyName + str(rows)].value
    stringFullLink = ws[ColumnGlassDoorFullLink + str(rows)].value
    PageNo = DecodePageNo(str(ws[ColumnPageNo + str(rows)].value))

    if(stringFullLink != None):
        stringFullLink = stringFullLink.strip()
    else:
        stringFullLink = ""
    
    if(companyNameFromSite != None):
        companyNameFromSite = companyNameFromSite.strip()
    else:
        companyNameFromSite = ""
    
        
    if(stringCompanyName != None):
        stringCompanyName = stringCompanyName.strip()
    else:
        stringCompanyName = ""
        
    print(ColumnCompanyName + str(rows) + ":Company Name:"+stringCompanyName)
    
    if(stringCompanyName != None and len(stringCompanyName) > 0):
        stringCompanyName = EncodeURL(stringCompanyName)
        #Query URL, parse and save
        URL = "https://www.glassdoor.com/Reviews/company-reviews.htm?suggestCount=0&suggestChosen=true&clickSource=searchBtn&typedKeyword={search}&sc.keyword={search}&locT=&locId=&jobType="
        URL = URL.replace("{search}", stringCompanyName)
        if( (len(companyNameFromSite) == 0 or len(stringFullLink) == 0) ):
            HTML = DownloadString(URL)
            companyNameFromSite = GetStringBetween(HTML, "<input name='sc.keyword' id='sc.keyword' class='keyword' type='text' tabindex='0' value='" , "'" , "" , True )
            IdFromSite = GetStringBetween(HTML, "<div class='vline cell showDesk'><i></i></div><a class='eiCell cell reviews'" , "><span class=" , "" , True )
            #href='/Reviews/Slalom-Reviews-E31102.htm'>
            IdFromSite = GetStringBetween(IdFromSite, "Reviews-", ".htm", "", True)
            if(len(IdFromSite) == 0):
                IdFromSite = GetStringBetween(HTML, "Reviews-", ".htm", "", True)
            
            link = GetStringBetween(HTML, "<a class='eiCell cell reviews' href='/" , ".htm'" , "", True )
            if(len(link) == 0):
                IdFromSite = ""
                stringFullLink = "!Not found in original site. Check URL in pageNo"
                PageNo = URL
            else:
                stringFullLink = "https://www.glassdoor.com/Reviews/" + link
                stringFullLink = stringFullLink + "_P{page}.htm"
                
            ws[ColumnGlassDoorID + str(rows)] =  IdFromSite
            ws[ColumnGlassDoorCompanyName + str(rows)] = companyNameFromSite
            ws[ColumnGlassDoorFullLink + str(rows)] = stringFullLink
            ws[ColumnPageNo + str(rows)] = str(PageNo)
            
            #Get the full link now
            PauseScript(5)
            if(True):
                print("Saving row: "+str(rows))
                wb.save(filenameExcel)

    if(fileobject != None):
            fileobject.close()
    fileobject = None


    #If we have a link then we create a filename and harvest?
    if(stringFullLink.find("https:", 0) == 0 and PageNo != -1):
        directoryPath = GetDirectoryPath(filenameExcel)
        import datetime
        csvpath = directoryPath + "\\" + companyNameFromSite + " " + datetime.datetime.now().strftime('%Y-%m-%d %H_%M_%S')+ ".csv"
                
        #For each page download strings
        beginURL = stringFullLink
                
        currenturl = beginURL
        currenturl = currenturl.replace("{page}", str(PageNo))
        if(PageNo == 1):
            currenturl = currenturl.replace("_P1", "")
                    
        #Harvest now
        HTML = DownloadString(currenturl)

        while(HTML != None):
                if(fileobject == None):
                    fileobject = WriteToFile(GetCSVHeader(), fileobject, csvpath)
                RowList = ParseHTML(HTML)
                for rowcount in range(0, len(RowList)):
                    RowItself = RowList[rowcount]
                    csvstring = ConvertToCSVString(RowItself)
                    fileobject = WriteToFile(csvstring, fileobject, csvpath)

                PageNo = PageNo + 1
                ws[ColumnPageNo + str(rows)] = str(PageNo)
                wb.save(filenameExcel)
                     
                currenturl = beginURL
                currenturl = currenturl.replace("{page}", str(PageNo))
                PauseScript(1) #how long between scrapes
                HTML = DownloadString(currenturl)
                if(HTML.find("<li class='next'><span class='disabled'><i><span>Next</span>", 0) != -1):
                    HTML = None
                    if(fileobject != None):
                        fileobject.close()
                    fileobject = None
                    PageNo=-1
                    ws[ColumnPageNo + str(rows)] = str(PageNo)
                    wb.save(filenameExcel)
    elif(len(stringCompanyName) == 0):
        break
wb.save(filenameExcel)
print("Done")
exit()
    









#end custom python scraper

    



